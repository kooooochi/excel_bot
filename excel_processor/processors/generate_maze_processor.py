"""サマリーシートを追加するプロセッサー"""

import random
import sys
from collections import deque
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_processor.base_processor import BaseSheetProcessor


class GenerateMazeProcessor(BaseSheetProcessor):
    """
    迷路を生成し、以下のシートを生成
    - Maze : 迷路
    - Distance : Start（S）からの距離
    - Path : Start（S）からGoal（G)までの最短距離

    設定例:
        height: 10
        width: 10
    """

    def process(self, workbook: Workbook, _file_path: str) -> Workbook:
        height = self.config.get("height", 10)
        width = self.config.get("width", 10)

        maze, start, goal = self._run_with_timer(
            "generate_maze",
            generate_maze,
            width=width,
            height=height,
        )

        visit = self._run_with_timer(
            "solver",
            solver,
            maze=maze,
            start=start,
            goal=goal,
        )

        self._run_with_timer(
            "output_maze_result",
            output_maze_result,
            workbook=workbook,
            maze=maze,
            visit=visit,
        )

        return workbook

    def _run_with_timer(self, process_name, function, *args, **kwargs):
        start_time = datetime.now()
        result = function(*args, **kwargs)
        end_time = datetime.now()
        elapsed_time = end_time - start_time
        print(
            f"[{self.__class__.__name__}][{process_name}] 処理時間: {elapsed_time}（{elapsed_time.total_seconds():.3f} 秒）"
        )
        return result


sys.setrecursionlimit(10**7)


def generate_maze(width, height):
    """
    width  : 迷路の横幅（奇数を推奨）
    height : 迷路の高さ（奇数を推奨）
    return : 迷路 2D 配列（壁=1, 道=0）、start座標、goal座標
    """

    _validate_maze_size(width, height)

    maze = [[1 for _ in range(width)] for _ in range(height)]

    # 開始位置（ランダムな奇数座標）
    start_x = random.randrange(1, width, 2)
    start_y = random.randrange(1, height, 2)
    maze[start_y][start_x] = 0

    # 移動方向
    directions = [(0, 2), (0, -2), (2, 0), (-2, 0)]

    def carve(x, y):
        random.shuffle(directions)
        for dx, dy in directions:
            nx, ny = x + dx, y + dy
            if 0 < nx < width-1 and 0 < ny < height-1:
                if maze[ny][nx] == 1:
                    maze[ny - dy // 2][nx - dx // 2] = 0
                    maze[ny][nx] = 0
                    carve(nx, ny)

    carve(start_x, start_y)

    # --------------------------------------
    # Start と Goal の決定
    # --------------------------------------

    # Start = 左上の最初の通路
    start_coord = None
    for y in range(height):
        for x in range(width):
            if maze[y][x] == 0:
                start_coord = (x, y)
                break
        if start_coord:
            break

    # Goal = 右下の最後の通路
    goal_coord = None
    for y in range(height - 1, -1, -1):
        for x in range(width - 1, -1, -1):
            if maze[y][x] == 0:
                goal_coord = (x, y)
                break
        if goal_coord:
            break

    return maze, start_coord, goal_coord


def _validate_maze_size(width: int, height: int):
    if width < 3 or height < 3:
        raise ValueError("Maze width and height must be >= 3.")
    if width % 2 == 0 or height % 2 == 0:
        raise ValueError("Maze width and height must be odd numbers.")


def print_maze(maze, start, goal):
    sx, sy = start
    gx, gy = goal

    for y, row in enumerate(maze):
        line = ""
        for x, cell in enumerate(row):
            if (x, y) == (sx, sy):
                line += "S"
            elif (x, y) == (gx, gy):
                line += "G"
            else:
                line += "#" if cell else " "
        print(line)




class State:
    def __init__(self, xy, cost):
        self._xy = xy
        self._cost = cost

    def get_position(self):
        return self._xy

    def get_cost(self):
        return self._cost

    def get_next_state(self, dx_dy):
        return State(
            (self._xy[0] + dx_dy[0], self._xy[1] + dx_dy[1]),
            self._cost + 1
        )



class Visit:
    def __init__(self, maze, start, goal):
        self._start = start
        self._goal = goal
        self._height = len(maze)
        self._width = len(maze[0])
        self._visit = [[-1 for _ in range(self._width)] for _ in range(self._height)]
        self._path = [[(-1, -1) for _ in range(self._width)] for _ in range(self._height)]
        self._maze = maze

    @property
    def start(self):
        return self._start

    @property
    def goal(self):
        return self._goal

    def get_cost(self, xy):
        return self._visit[xy[1]][xy[0]]

    def set_cost(self, state):
        xy = state.get_position()
        cost = state.get_cost()
        if not self.can_move(xy):
            raise ValueError(f"Invalid position {xy}. Out of bounds or wall.")
        self._visit[xy[1]][xy[0]] = cost

    def set_path(self, now_state, next_state):
        now_xy = now_state.get_position()
        next_xy = next_state.get_position()
        self._path[next_xy[1]][next_xy[0]] = now_xy

    def update(self, now_state, next_state):
        self.set_cost(next_state)
        self.set_path(now_state, next_state)

    def get_start_to_goal_path(self):
        path = []
        goal = tuple(self._goal)
        start = tuple(self._start)

        # 目標が未到達
        if self._path[goal[1]][goal[0]] == (-1, -1) and goal != start:
            return path

        now = goal
        while now != (-1, -1):
            path.append(now)
            if now == start:
                break
            now = self._path[now[1]][now[0]]

        return list(reversed(path))
        
    def can_move(self, xy):
        return (
            0 <= xy[1] < self._height
            and 0 <= xy[0] < self._width
            and self._maze[xy[1]][xy[0]] != 1
        )

    def print_visit(self):
        for value in self._visit:
            print(value)

    def try_move(self, now_state, next_state):
        """
        now_state から next_state への遷移が有効で未訪問なら更新する
        戻り値: True=更新した, False=スキップ
        """
        next_xy = next_state.get_position()
        if (not self.can_move(next_xy)) or self.get_cost(next_xy) >= 0:
            return False
        self.update(now_state, next_state)
        return True

    def get_visit_map(self):
        """訪問コストの2次元配列を返す（内部参照をそのまま返すので編集しないこと）"""
        return self._visit
            


def solver(maze, start, goal):
    visit = Visit(maze, start, goal)
    directions = [(1, 0), (-1, 0), (0, 1), (0, -1)]
    
    queue = deque()
    start_state = State(start, 0)
    visit.set_cost(start_state)
    queue.append(start_state)
    while queue:
        now_state = queue.popleft()

        for dx_dy in directions:
            next_state = now_state.get_next_state(dx_dy)
            if visit.try_move(now_state, next_state):
                queue.append(next_state)
    
    return visit


def output_maze_result(workbook, maze, visit):
    """
    迷路、距離マップ、最短経路をそれぞれ別シートに出力する
    """
    sheet_names = ("Maze", "Distance", "Path")

    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    wall_fill = PatternFill(fill_type="solid", fgColor="404040")
    start_fill = PatternFill(fill_type="solid", fgColor="4CAF50")
    goal_fill = PatternFill(fill_type="solid", fgColor="F44336")
    path_fill = PatternFill(fill_type="solid", fgColor="FFD54F")
    text_center = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    neutral_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    num_font = Font(color="0F172A")
    cell_size = 3  # おおよそ正方形に見える幅・高さ（単位: Excel の列幅/行高さ単位）

    # Maze シート
    maze_ws = workbook.create_sheet("Maze")
    maze_ws.sheet_view.showGridLines = False
    for y, row in enumerate(maze):
        for x, cell in enumerate(row):
            c = maze_ws.cell(row=y + 1, column=x + 1)
            if (x, y) == visit.start:
                c.value = "S"
                c.fill = start_fill
                c.font = bold_font
            elif (x, y) == visit.goal:
                c.value = "G"
                c.fill = goal_fill
                c.font = bold_font
            else:
                c.value = ""
                c.fill = wall_fill if cell else neutral_fill
            c.alignment = text_center
    for col in range(1, len(maze[0]) + 1):
        maze_ws.column_dimensions[get_column_letter(col)].width = cell_size
    for row_idx in range(1, len(maze) + 1):
        maze_ws.row_dimensions[row_idx].height = cell_size * 5  # 行高さは幅より大きめ係数

    # Distance シート
    dist_ws = workbook.create_sheet("Distance")
    dist_ws.sheet_view.showGridLines = False
    visit_map = visit.get_visit_map()
    max_cost = max(max(row) for row in visit_map if row) or 0
    for y, row in enumerate(visit_map):
        for x, cost in enumerate(row):
            c = dist_ws.cell(row=y + 1, column=x + 1, value=cost)
            c.alignment = text_center
            c.font = num_font
            if cost >= 0 and max_cost > 0:
                # 簡易ヒートマップ: コストに応じて薄い青から濃い青へ
                intensity = int(255 - (cost / max_cost) * 120)
                hex_part = f"{intensity:02X}"
                c.fill = PatternFill(fill_type="solid", fgColor=f"BB{hex_part}FF")
            else:
                c.fill = wall_fill
    for col in range(1, len(maze[0]) + 1):
        dist_ws.column_dimensions[get_column_letter(col)].width = cell_size
    for row_idx in range(1, len(maze) + 1):
        dist_ws.row_dimensions[row_idx].height = cell_size * 5

    # Path シート
    path_ws = workbook.create_sheet("Path")
    path_ws.sheet_view.showGridLines = False
    path = visit.get_start_to_goal_path()
    step_lookup = {xy: idx for idx, xy in enumerate(path)}
    for y, row in enumerate(maze):
        for x, cell in enumerate(row):
            coord = (x, y)
            c = path_ws.cell(row=y + 1, column=x + 1)
            if coord == visit.start:
                c.value = "S"
                c.fill = start_fill
                c.font = bold_font
            elif coord == visit.goal:
                c.value = "G"
                c.fill = goal_fill
                c.font = bold_font
            elif coord in step_lookup:
                c.value = step_lookup[coord]
                c.fill = path_fill
                c.font = num_font
            elif cell == 1:
                c.value = ""
                c.fill = wall_fill
            else:
                c.value = ""
                c.fill = neutral_fill
            c.alignment = text_center
    for col in range(1, len(maze[0]) + 1):
        path_ws.column_dimensions[get_column_letter(col)].width = cell_size
    for row_idx in range(1, len(maze) + 1):
        path_ws.row_dimensions[row_idx].height = cell_size * 5
