"""サマリーシートを追加するプロセッサー"""

from datetime import datetime
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from excel_processor.base_processor import BaseSheetProcessor


class SummarySheetProcessor(BaseSheetProcessor):
    """
    サマリーシートを追加するプロセッサー

    設定例:
        sheet_name: "Summary"  # シート名（デフォルト: "Summary"）
        position: 0  # シートの位置（0=先頭、デフォルト: 0）
    """

    def process(self, workbook: Workbook, file_path: str) -> Workbook:
        sheet_name = self.config.get('sheet_name', 'Summary')
        position = self.config.get('position', 0)

        self.log(f"Adding summary sheet: {sheet_name}")

        # サマリーシートを作成
        if sheet_name in workbook.sheetnames:
            # 既存のシートを削除
            del workbook[sheet_name]

        summary_sheet = workbook.create_sheet(sheet_name, position)

        # ヘッダー
        summary_sheet['A1'] = "Excel Processing Summary"
        summary_sheet['A1'].font = Font(size=16, bold=True)

        # 基本情報
        row = 3
        summary_sheet[f'A{row}'] = "Processing Date:"
        summary_sheet[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row += 1
        summary_sheet[f'A{row}'] = "File Name:"
        summary_sheet[f'B{row}'] = file_path

        row += 1
        summary_sheet[f'A{row}'] = "Total Sheets:"
        summary_sheet[f'B{row}'] = len(workbook.sheetnames) - 1  # サマリーシート自体を除く

        # シートリスト
        row += 2
        summary_sheet[f'A{row}'] = "Sheet List:"
        summary_sheet[f'A{row}'].font = Font(bold=True)

        row += 1
        for idx, sheet_name in enumerate(workbook.sheetnames, 1):
            if sheet_name != self.config.get('sheet_name', 'Summary'):
                summary_sheet[f'A{row}'] = f"{idx - 1}. {sheet_name}"

                # シートの行数と列数を取得
                ws = workbook[sheet_name]
                summary_sheet[f'B{row}'] = f"Rows: {ws.max_row}, Cols: {ws.max_column}"
                row += 1

        # 列幅を調整
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 40

        self.log("Summary sheet added successfully")
        return workbook
