"""書式を適用するプロセッサー"""

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from excel_processor.base_processor import BaseSheetProcessor


class FormatProcessor(BaseSheetProcessor):
    """
    全シートにフォーマットを適用するプロセッサー

    設定例:
        header_color: "4472C4"  # ヘッダー背景色（16進数）
        font_name: "Arial"  # フォント名
        font_size: 11  # フォントサイズ
        apply_borders: true  # 罫線を適用するか
        auto_width: true  # 列幅を自動調整するか
        exclude_sheets: ["Summary"]  # 除外するシート名
    """

    def process(self, workbook: Workbook, file_path: str) -> Workbook:
        header_color = self.config.get('header_color', '4472C4')
        font_name = self.config.get('font_name', 'Arial')
        font_size = self.config.get('font_size', 11)
        apply_borders = self.config.get('apply_borders', True)
        auto_width = self.config.get('auto_width', True)
        exclude_sheets = self.config.get('exclude_sheets', [])

        self.log("Applying formatting to all sheets")

        for sheet_name in workbook.sheetnames:
            if sheet_name in exclude_sheets:
                self.log(f"Skipping sheet: {sheet_name}")
                continue

            ws = workbook[sheet_name]
            self.log(f"Formatting sheet: {sheet_name}")

            # ヘッダー行（1行目）のフォーマット
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = Font(name=font_name, size=font_size, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # データ行のフォント設定
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.font.size is None or cell.font.name is None:
                        cell.font = Font(name=font_name, size=font_size)

            # 罫線を適用
            if apply_borders:
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border

            # 列幅の自動調整
            if auto_width:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        self.log("Formatting completed")
        return workbook
