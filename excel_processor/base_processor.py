"""ベースプロセッサー - ユーザーがカスタマイズ可能な処理インターフェース"""

from abc import ABC, abstractmethod
from typing import Any, Dict
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class BaseSheetProcessor(ABC):
    """
    シート処理のベースクラス

    ユーザーはこのクラスを継承して、カスタム処理を実装します。
    """

    def __init__(self, config: Dict[str, Any] = None):
        """
        Args:
            config: プロセッサーの設定（YAML設定ファイルから読み込まれる）
        """
        self.config = config or {}

    @abstractmethod
    def process(self, workbook: Workbook, file_path: str) -> Workbook:
        """
        Excelファイルを処理するメインメソッド

        Args:
            workbook: openpyxlのWorkbookオブジェクト
            file_path: 処理中のファイルパス（参照用）

        Returns:
            処理済みのWorkbookオブジェクト
        """
        pass

    def create_sheet(self, workbook: Workbook, sheet_name: str, index: int = None) -> Worksheet:
        """
        新しいシートを作成するヘルパーメソッド

        Args:
            workbook: Workbookオブジェクト
            sheet_name: 新しいシート名
            index: シートの挿入位置（Noneの場合は最後に追加）

        Returns:
            作成されたWorksheetオブジェクト
        """
        if index is not None:
            return workbook.create_sheet(sheet_name, index)
        return workbook.create_sheet(sheet_name)

    def get_or_create_sheet(self, workbook: Workbook, sheet_name: str) -> Worksheet:
        """
        シートを取得、存在しない場合は作成

        Args:
            workbook: Workbookオブジェクト
            sheet_name: シート名

        Returns:
            Worksheetオブジェクト
        """
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        return workbook.create_sheet(sheet_name)

    def log(self, message: str):
        """ログ出力用ヘルパーメソッド"""
        print(f"[{self.__class__.__name__}] {message}")
