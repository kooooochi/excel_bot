"""ヘルパー関数とユーティリティ"""

from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.workbook import Workbook


def load_excel_from_input(
    file_name: Optional[str] = None,
    input_dir: str = "input"
) -> tuple[Workbook, Path]:
    """
    inputディレクトリからExcelファイルを読み込む

    Args:
        file_name: ファイル名（Noneの場合は最初のファイル）
        input_dir: 入力ディレクトリ

    Returns:
        (Workbook, Path): ワークブックオブジェクトとファイルパス

    Raises:
        FileNotFoundError: ファイルが見つからない場合
        ValueError: Excelファイルが存在しない場合
    """
    input_path = Path(input_dir)

    if not input_path.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    # Excelファイル一覧を取得
    excel_files = get_excel_files(input_dir)

    if not excel_files:
        raise ValueError(f"No Excel files found in {input_dir}")

    # ファイルを選択
    if file_name:
        target_file = input_path / file_name
        if not target_file.exists():
            raise FileNotFoundError(f"File not found: {target_file}")
    else:
        target_file = excel_files[0]
        print(f"Loading first file: {target_file.name}")

    # ワークブックを読み込み
    workbook = openpyxl.load_workbook(target_file)
    print(f"Loaded: {target_file.name}")
    print(f"Sheets: {workbook.sheetnames}")

    return workbook, target_file


def get_excel_files(input_dir: str = "input") -> List[Path]:
    """
    inputディレクトリからExcelファイルのリストを取得

    Args:
        input_dir: 入力ディレクトリ

    Returns:
        Excelファイルのパスのリスト
    """
    input_path = Path(input_dir)

    if not input_path.exists():
        return []

    excel_files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))

    # 一時ファイルを除外
    excel_files = [f for f in excel_files if not f.name.startswith("~$")]

    return sorted(excel_files)


def save_preview(
    workbook: Workbook,
    original_file: Path,
    preview_dir: str = "output/preview"
) -> Path:
    """
    処理結果をプレビュー用に保存

    Args:
        workbook: 保存するワークブック
        original_file: 元のファイルパス
        preview_dir: プレビュー保存先ディレクトリ

    Returns:
        保存したファイルのパス
    """
    preview_path = Path(preview_dir)
    preview_path.mkdir(parents=True, exist_ok=True)

    preview_file = preview_path / f"preview_{original_file.name}"
    workbook.save(preview_file)

    print(f"Preview saved: {preview_file}")
    return preview_file


def print_sheet_info(workbook: Workbook):
    """
    ワークブックのシート情報を出力

    Args:
        workbook: ワークブック
    """
    print(f"\nTotal sheets: {len(workbook.sheetnames)}")
    print(f"Active sheet: {workbook.active.title}")
    print("\nSheet details:")

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"  - {sheet_name}: {ws.max_row} rows x {ws.max_column} cols")


def print_sheet_preview(
    workbook: Workbook,
    sheet_name: Optional[str] = None,
    max_rows: int = 10
):
    """
    シートの内容をプレビュー表示

    Args:
        workbook: ワークブック
        sheet_name: シート名（Noneの場合はアクティブシート）
        max_rows: 表示する最大行数
    """
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found")
            return
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    print(f"\nSheet: {ws.title}")
    print(f"Size: {ws.max_row} rows x {ws.max_column} cols")
    print(f"\nFirst {max_rows} rows:")

    for row in ws.iter_rows(
        min_row=1,
        max_row=min(max_rows, ws.max_row),
        values_only=True
    ):
        print(row)
